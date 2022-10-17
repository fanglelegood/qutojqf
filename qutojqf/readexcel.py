    # This Python file uses the following encoding: utf-8

# if __name__ == "__main__":
#     pass
import openpyxl
from win32com.client import Dispatch

def save_to_file(file_name, contents):
    fh = open(file_name, 'w')
    fh.write(contents)
    fh.close()

def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

def getbladedata(seriesnumber,filename):
    book1 = openpyxl.load_workbook(filename, data_only=False)
    ws1 = book1.get_sheet_by_name(u"叶片数据输入")
    ws1['B2'] = seriesnumber
    book1.save(filename)
    just_open(filename)
    book2 = openpyxl.load_workbook(filename, data_only=True)
    ws2 = book2.get_sheet_by_name(u"首导叶绘图表")
    ws3 = book2.get_sheet_by_name(u"首导叶绘图表")
    ws4 = book2.get_sheet_by_name(u"末导叶绘图表")
    ws5 = book2.get_sheet_by_name(u"末导叶根绘图表")
    ws6 = book2.get_sheet_by_name(u"动叶片绘图表")
    sbladedata1=[]
    sbladedata2=[]
    sbladedata3=[]
    sbladedata4=[]
    mbladedata=[]
    for row in ws6['C3:C44']:
        for cell in row:
            mbladedata.append(str(cell.value))
    for row in ws2['C3:C39']:
        for cell in row:
            sbladedata1.append(str(cell.value))
    for row in ws3['C3:C35']:
        for cell in row:
            sbladedata2.append(str(cell.value))
    for row in ws4['C3:C36']:
        for cell in row:
            sbladedata3.append(str(cell.value))
    for row in ws5['C3:C19']:
        for cell in row:
            sbladedata4.append(str(cell.value))
    return mbladedata,sbladedata1,sbladedata2,sbladedata3,sbladedata4

# (mbladedata,sbladedata1,sbladedata2,sbladedata3,sbladedata4)= getbladedata(1)
# mbladedatastring = '(command "GATTE" "b" "动叶12.5-56.7" "A" "' + mbladedata[0] + '" "Y")'  \
#                     '(command "GATTE" "b" "动叶12.5-56.7" "AG0" "' + mbladedata[1] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "AB0" "' + mbladedata[2] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "ALA" "' + mbladedata[3] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "AB" "' + mbladedata[4] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "AB1" "' + mbladedata[5] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HLA" "' + mbladedata[6] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HLA0" "' + mbladedata[7] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HG0" "' + mbladedata[8] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HGA" "' + mbladedata[9] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HA" "' + mbladedata[10] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "HGA1" "' + mbladedata[11] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RG" "' + mbladedata[12] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "DA1" "' + mbladedata[13] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RLE1" "' + mbladedata[14] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RLE3" "' + mbladedata[15] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RLE2" "' + mbladedata[16] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RLE4" "' + mbladedata[17] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "DA" "' + mbladedata[18] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "CF" "' + mbladedata[19] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "AM" "' + mbladedata[20] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "H3" "' + mbladedata[21] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "DQ" "' + mbladedata[22] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "安装角B" "' + mbladedata[23] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "L" "' + mbladedata[24] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "LP1" "' + mbladedata[25] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "LP2" "' + mbladedata[26] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "YH" "' + mbladedata[27] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "RG0" "' + mbladedata[28] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "P" "' + mbladedata[29] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "Q" "' + mbladedata[30] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "菱形角" "' + mbladedata[31] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "B" "' + mbladedata[32] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "YS" "' + mbladedata[33] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "YD" "' + mbladedata[34] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "叶型号" "' + mbladedata[35] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "叶根轮槽图号" "' + mbladedata[36] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "PH" "' + mbladedata[37] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "QH" "' + mbladedata[38] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "ZA" "' + mbladedata[39] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "ZA1" "' + mbladedata[40] + '" "Y")'\
#                     '(command "GATTE" "b" "动叶12.5-56.7" "ZH" "' + mbladedata[41] + '" "Y")'
# print(mbladedatastring)
# save_to_file('D:/qutojqf/new/Bladedata/mbladedatastring.txt',mbladedatastring)


def GenerateTxt(series,filename):
    i=1
    while i<series+1:
        (mbladedata,sbladedata1,sbladedata2,sbladedata3,sbladedata4)= getbladedata(i,filename)
        mbladedatastring = '(command "GATTE" "b" "动叶12.5-56.7" "A" "' + mbladedata[0] + '" "Y")'  \
                    '(command "GATTE" "b" "动叶12.5-56.7" "AG0" "' + mbladedata[1] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "AB0" "' + mbladedata[2] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "ALA" "' + mbladedata[3] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "AB" "' + mbladedata[4] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "AB1" "' + mbladedata[5] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HLA" "' + mbladedata[6] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HLA0" "' + mbladedata[7] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HG0" "' + mbladedata[8] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HGA" "' + mbladedata[9] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HA" "' + mbladedata[10] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "HGA1" "' + mbladedata[11] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RG" "' + mbladedata[12] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "DA1" "' + mbladedata[13] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RLE1" "' + mbladedata[14] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RLE3" "' + mbladedata[15] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RLE2" "' + mbladedata[16] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RLE4" "' + mbladedata[17] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "DA" "' + mbladedata[18] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "CF" "' + mbladedata[19] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "AM" "' + mbladedata[20] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "H3" "' + mbladedata[21] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "DQ" "' + mbladedata[22] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "安装角B" "' + mbladedata[23] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "L" "' + mbladedata[24] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "LP1" "' + mbladedata[25] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "LP2" "' + mbladedata[26] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "YH" "' + mbladedata[27] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "RG0" "' + mbladedata[28] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "P" "' + mbladedata[29] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "Q" "' + mbladedata[30] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "菱形角" "' + mbladedata[31] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "B" "' + mbladedata[32] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "YS" "' + mbladedata[33] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "YD" "' + mbladedata[34] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "叶型号" "' + mbladedata[35] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "叶根轮槽图号" "' + mbladedata[36] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "PH" "' + mbladedata[37] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "QH" "' + mbladedata[38] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "ZA" "' + mbladedata[39] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "ZA1" "' + mbladedata[40] + '" "Y")'\
                    '(command "GATTE" "b" "动叶12.5-56.7" "ZH" "' + mbladedata[41] + '" "Y")'
        save_to_file('D:/qutojqf/new/Bladedata/'+str(i)+'-d.txt',mbladedatastring)
        sbladedata1string = '(command "GATTE" "b" "导叶12.5-54.7" "HLEJ" "' + sbladedata1[0] + '" "Y")'  \
                    '(command "GATTE" "b" "导叶12.5-54.7" "AG" "' + sbladedata1[1] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HG1" "' + sbladedata1[2] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "H" "' + sbladedata1[3] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HG2" "' + sbladedata1[4] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HG" "' + sbladedata1[5] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HLE" "' + sbladedata1[6] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "A" "' + sbladedata1[7] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "D1" "' + sbladedata1[8] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "ALE" "' + sbladedata1[9] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "AB" "' + sbladedata1[10] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RG1" "' + sbladedata1[11] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RG2" "' + sbladedata1[12] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RG3" "' + sbladedata1[13] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "Δ" "' + sbladedata1[14] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HG3" "' + sbladedata1[15] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "HG4" "' + sbladedata1[16] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RLE1" "' + sbladedata1[17] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RLE2" "' + sbladedata1[18] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RLE3" "' + sbladedata1[19] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "RLE4" "' + sbladedata1[20] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "YS" "' + sbladedata1[21] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "YD" "' + sbladedata1[22] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "菱形角" "' + sbladedata1[23] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "B" "' + sbladedata1[24] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "Q" "' + sbladedata1[25] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "P" "' + sbladedata1[26] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "LP1" "' + sbladedata1[27] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "LP2" "' + sbladedata1[28] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "L" "' + sbladedata1[29] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "安装角B" "' + sbladedata1[30] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "YH" "' + sbladedata1[31] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "叶型号" "' + sbladedata1[32] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "叶片数" "' + sbladedata1[33] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "厚叶" "' + sbladedata1[34] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "PH" "' + sbladedata1[35] + '" "Y")'\
                    '(command "GATTE" "b" "导叶12.5-54.7" "QH" "' + sbladedata1[36] + '" "Y")'
        save_to_file('D:/qutojqf/new/Bladedata/'+str(i)+'-j1.txt',sbladedata1string)
        sbladedata2string = '(command "GATTE" "b" "首导叶12.5-54.7" "SQ" "' + sbladedata2[0] + '" "Y")'  \
                    '(command "GATTE" "b" "首导叶12.5-54.7" "SP" "' + sbladedata2[1] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "叶型号" "' + sbladedata2[2] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "YH" "' + sbladedata2[3] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "安装角B" "' + sbladedata2[4] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "L" "' + sbladedata2[5] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "LP2" "' + sbladedata2[6] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "LP1" "' + sbladedata2[7] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "B" "' + sbladedata2[8] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "菱形角" "' + sbladedata2[9] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "YD" "' + sbladedata2[10] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "YS" "' + sbladedata2[11] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RLE4" "' + sbladedata2[12] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RLE3" "' + sbladedata2[13] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RLE2" "' + sbladedata2[14] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RLE1" "' + sbladedata2[15] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HG4" "' + sbladedata2[16] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HG3" "' + sbladedata2[17] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "Δ" "' + sbladedata2[18] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RG3" "' + sbladedata2[19] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RG2" "' + sbladedata2[20] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "RG1" "' + sbladedata2[21] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "AB" "' + sbladedata2[22] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "ALE" "' + sbladedata2[23] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "D1" "' + sbladedata2[24] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "A" "' + sbladedata2[25] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HLE" "' + sbladedata2[26] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HG" "' + sbladedata2[27] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HG2" "' + sbladedata2[28] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "H" "' + sbladedata2[29] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HG1" "' + sbladedata2[30] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "AG" "' + sbladedata2[31] + '" "Y")'\
                    '(command "GATTE" "b" "首导叶12.5-54.7" "HLEJ" "' + sbladedata2[32] + '" "Y")'
        save_to_file('D:/qutojqf/new/Bladedata/'+str(i)+'-j2.txt',sbladedata2string)
        sbladedata3string = '(command "GATTE" "b" "末导叶12.5-54.7" "MQ" "' + sbladedata3[0] + '" "Y")'  \
                    '(command "GATTE" "b" "末导叶12.5-54.7" "叶型号" "' + sbladedata3[1] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "YH" "' + sbladedata3[2] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "安装角B" "' + sbladedata3[3] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "L" "' + sbladedata3[4] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "LP2" "' + sbladedata3[5] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "LP1" "' + sbladedata3[6] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "P" "' + sbladedata3[7] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "Q" "' + sbladedata3[8] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "B" "' + sbladedata3[9] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "菱形角" "' + sbladedata3[10] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "YD" "' + sbladedata3[11] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "YS" "' + sbladedata3[12] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RLE4" "' + sbladedata3[13] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RLE3" "' + sbladedata3[14] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RLE2" "' + sbladedata3[15] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RLE1" "' + sbladedata3[16] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG4" "' + sbladedata3[17] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG3" "' + sbladedata3[18] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "Δ" "' + sbladedata3[19] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG3" "' + sbladedata3[20] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG2" "' + sbladedata3[21] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG1" "' + sbladedata3[22] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "AB" "' + sbladedata3[23] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "ALE" "' + sbladedata3[24] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "D1" "' + sbladedata3[25] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "A" "' + sbladedata3[26] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HLE" "' + sbladedata3[27] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG" "' + sbladedata3[28] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG2" "' + sbladedata3[29] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "H" "' + sbladedata3[30] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG1" "' + sbladedata3[31] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "AG" "' + sbladedata3[32] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HLEJ" "' + sbladedata3[33] + '" "Y")'
        save_to_file('D:/qutojqf/new/Bladedata/'+str(i)+'-j3.txt',sbladedata3string)
        sbladedata4string = '(command "GATTE" "b" "末导叶12.5-54.7" "MH" "' + sbladedata4[0] + '" "Y")'  \
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HGJ" "' + sbladedata4[1] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "AG" "' + sbladedata4[2] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG1" "' + sbladedata4[3] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG2" "' + sbladedata4[4] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG" "' + sbladedata4[5] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "A" "' + sbladedata4[6] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "ALE" "' + sbladedata4[7] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "AB" "' + sbladedata4[8] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG1" "' + sbladedata4[9] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG2" "' + sbladedata4[10] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "RG3" "' + sbladedata4[11] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "Δ" "' + sbladedata4[12] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG3" "' + sbladedata4[13] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "HG4" "' + sbladedata4[14] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "菱形角" "' + sbladedata4[15] + '" "Y")'\
                    '(command "GATTE" "b" "末导叶12.5-54.7" "B" "' + sbladedata4[16] + '" "Y")'                    
        save_to_file('D:/qutojqf/new/Bladedata/'+str(i)+'-j4.txt',sbladedata4string)
        i =  i+1
GenerateTxt(2,'D:/qutojqf/new/HS23119JQFbladedrawingtable.xlsx')


# (sbladedata1,sbladedata2,sbladedata3,sbladedata4,mbladedata)= getbladedata(22)
# print(sbladedata1)
# print(sbladedata2)
# print(sbladedata3)
# print(sbladedata4)
# print(mbladedata)






#c6_value = ws6.cell(3, 3).value
#c6_value = ws6['C3':'C45']
#print(c6_value)
#for row in ws6['C3:C44']:
#    for cell in row:
#        print(cell.value)


